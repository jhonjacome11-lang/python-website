from flask import Flask, render_template, request, jsonify, Response, send_file, session
import json, os, io
from datetime import datetime
from functools import wraps

app = Flask(__name__)
app.secret_key    = os.environ.get('SECRET_KEY', 'purochancho-dev-key-2026')
ADMIN_PASSWORD    = os.environ.get('ADMIN_PASSWORD', '')
WA_NUMBER         = os.environ.get('WA_NUMBER', '')

# ── DECORADOR DE AUTENTICACIÓN ────────────────────────────────
def require_admin(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin'):
            return jsonify({'error': 'No autorizado'}), 401
        return f(*args, **kwargs)
    return decorated

# ── HELPERS JSON ───────────────────────────────────────────────
def _ensure_data():
    os.makedirs('data', exist_ok=True)

def _read(path, default):
    _ensure_data()
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return default

def _write(path, data):
    _ensure_data()
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# ── PÁGINA PRINCIPAL ───────────────────────────────────────────────
@app.route('/')
def home():
    return Response(render_template('home.html'), mimetype='text/html')

# ── CONFIG PÚBLICA ───────────────────────────────────────────────
@app.route('/api/config')
def get_config():
    return jsonify({'wa': WA_NUMBER})

# ── AUTH ──────────────────────────────────────────────────────
@app.route('/api/login', methods=['POST'])
def login_admin():
    d  = request.get_json() or {}
    pw = d.get('password', '')
    if ADMIN_PASSWORD and pw == ADMIN_PASSWORD:
        session.permanent = True
        session['admin']  = True
        return jsonify({'success': True})
    return jsonify({'success': False})

@app.route('/api/logout', methods=['POST'])
def logout_admin():
    session.clear()
    return jsonify({'ok': True})

@app.route('/api/check-auth')
def check_auth():
    return jsonify({'authenticated': bool(session.get('admin'))})

# ── CLIENTES ──────────────────────────────────────────────────
@app.route('/api/cliente', methods=['POST'])
def guardar_cliente():
    d = request.get_json() or {}
    clientes = _read('data/clientes.json', [])
    nuevo = {
        'id':       len(clientes) + 1,
        'nombre':   d.get('nombre', ''),
        'telefono': d.get('telefono', ''),
        'sector':   d.get('sector', ''),
        'producto': d.get('producto', ''),
        'nota':     d.get('nota', ''),
        'fecha':    datetime.now().strftime('%Y-%m-%d %H:%M'),
    }
    clientes.append(nuevo)
    _write('data/clientes.json', clientes)
    return jsonify({'ok': True, 'id': nuevo['id']})

@app.route('/api/clientes', methods=['GET'])
@require_admin
def obtener_clientes():
    return jsonify(_read('data/clientes.json', []))

@app.route('/api/cliente/<int:cid>', methods=['DELETE'])
@require_admin
def eliminar_cliente(cid):
    clientes = _read('data/clientes.json', [])
    clientes = [c for c in clientes if c.get('id') != cid]
    _write('data/clientes.json', clientes)
    return jsonify({'ok': True})

# ── PRECIOS ───────────────────────────────────────────────────
DEFAULT_PRICES = {
    'lechon':      {'min': 70, 'max': 110},
    'cerdo_pie':   {'precio_lb': 2.50},
    'cortes':      {'precio_lb': 2.50},
    'cuts_status': {
        'pierna':'disponible','lomo':'disponible','costilla':'disponible',
        'panceta':'disponible','paleta':'disponible','cabeza':'disponible',
    },
}

DEFAULT_INVENTORY = {
    'lechones':   {'cantidad': 4, 'reservados': 0, 'proxima': ''},
    'cerdo_pie':  {'cantidad': 2, 'reservados': 0, 'proxima': ''},
    'cortes':     {'estado': 'disponible', 'proxima_faena': ''},
}

DEFAULT_CONTENT = {
    'hero_titulo':    'Del criadero\na tu mesa.',
    'hero_subtitulo': 'Lechones, cerdos en pie y cortes frescos con atención directa por WhatsApp.',
    'horario':        'Lunes a sábado · 7:00 AM – 6:00 PM',
    'facebook':       '@PuroChanchoFamiliar',
    'instagram':      '@PuroChanchoFamiliar',
}

@app.route('/api/precios', methods=['GET'])
def obtener_precios():
    return jsonify(_read('data/precios.json', DEFAULT_PRICES))

@app.route('/api/precios', methods=['POST'])
@require_admin
def guardar_precios():
    data = request.get_json() or {}
    current = _read('data/precios.json', DEFAULT_PRICES)
    current.update(data)
    _write('data/precios.json', current)
    return jsonify({'ok': True})

@app.route('/api/inventory', methods=['GET'])
def obtener_inventario():
    return jsonify(_read('data/inventory.json', DEFAULT_INVENTORY))

@app.route('/api/inventory', methods=['POST'])
@require_admin
def guardar_inventario():
    data = request.get_json() or {}
    _write('data/inventory.json', data)
    return jsonify({'ok': True})

@app.route('/api/content', methods=['GET'])
def obtener_content():
    return jsonify(_read('data/content.json', DEFAULT_CONTENT))

@app.route('/api/content', methods=['POST'])
@require_admin
def guardar_content():
    data = request.get_json() or {}
    _write('data/content.json', data)
    return jsonify({'ok': True})

# ── ESTADÍSTICAS REALES ────────────────────────────────────────
@app.route('/api/stats')
@require_admin
def get_stats():
    clientes  = _read('data/clientes.json', [])
    total     = len(clientes)
    lechones  = sum(1 for c in clientes if 'echon'  in (c.get('producto') or '').lower())
    cerdo_pie = sum(1 for c in clientes if 'pie'    in (c.get('producto') or '').lower())
    cortes_n  = sum(1 for c in clientes if 'orte'   in (c.get('producto') or '').lower())

    # Pedidos por mes (últimos 6 meses)
    meses = {}
    for c in clientes:
        try:
            mes = c.get('fecha', '')[:7]  # '2026-04'
            meses[mes] = meses.get(mes, 0) + 1
        except Exception:
            pass
    meses_sorted = sorted(meses.items())[-6:]

    ultimos = sorted(clientes, key=lambda x: x.get('fecha',''), reverse=True)[:5]

    return jsonify({
        'total': total,
        'lechones': lechones,
        'cerdo_pie': cerdo_pie,
        'cortes': cortes_n,
        'por_mes': meses_sorted,
        'ultimos': ultimos,
    })

# ── EXPORTAR EXCEL (protegido) ───────────────────────────────────
@app.route('/api/export/excel')
@require_admin
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({'error': 'openpyxl no instalado'}), 500

    clientes = _read('data/clientes.json', [])
    wb = openpyxl.Workbook()

    header_fill = PatternFill('solid', fgColor='1C1A16')
    header_font = Font(color='FBF8F2', bold=True)

    categorias = [
        ('Todos',       clientes),
        ('Lechones',    [c for c in clientes if 'echon' in (c.get('producto') or '')]),
        ('Cerdo en pie',[c for c in clientes if 'pie'   in (c.get('producto') or '').lower()]),
        ('Cortes',      [c for c in clientes if 'orte'  in (c.get('producto') or '')]),
    ]

    cols = ['ID', 'Nombre', 'Teléfono', 'Sector', 'Producto', 'Nota', 'Fecha']
    keys = ['id',  'nombre', 'telefono', 'sector', 'producto', 'nota', 'fecha']

    for i, (titulo, filas) in enumerate(categorias):
        ws = wb.active if i == 0 else wb.create_sheet(titulo)
        ws.title = titulo
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        for ri, cliente in enumerate(filas, 2):
            for ci, key in enumerate(keys, 1):
                ws.cell(row=ri, column=ci, value=str(cliente.get(key, '')))
        for col in ws.columns:
            ancho = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(ancho + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre = f'puro_chancho_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    return send_file(buf, as_attachment=True, download_name=nombre,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    app.run(debug=True)
